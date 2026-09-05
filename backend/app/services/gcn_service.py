from __future__ import annotations

import io
import os
from datetime import datetime, timezone

import requests
from flask import current_app, jsonify
from openpyxl import Workbook, load_workbook

import gcn_sync
from ..repositories import gcn_repository, supabase_client
from . import parcel_service

# Nhóm KH 2959 hợp lệ cho chức năng "Xuất GCN theo nhóm".
_NHOM_HOP_LE = {"NHÓM 1", "NHÓM 2"}
_EXPORT_PAGE = 5000      # số dòng mỗi lần gọi RPC
_EXPORT_MAX = 300_000    # trần an toàn cho 1 lần xuất

# File mẫu TANAN.xlsx (biểu tổng hợp dữ liệu GCN) — hàng 1-4 là tiêu đề
# (có gộp ô), dữ liệu mẫu ở hàng 5 trở đi bị xóa và thay bằng dữ liệu thật
# khi xuất. Đường dẫn tương đối tới backend/data/ (xem backend/data/Thua_Dat.xsd
# cho cùng kiểu file tĩnh đóng gói theo backend).
_MAU_TANAN_PATH = os.path.join(
    os.path.dirname(__file__), "..", "..", "data", "TANAN.xlsx"
)
_MAU_TANAN_HEADER_ROWS = 4  # hàng 1-4: tiêu đề; dữ liệu bắt đầu từ hàng 5
_MAU_TANAN_SHEET = "Sheet1"

# Cột B..BG của mẫu TANAN.xlsx, ánh xạ theo TÊN cột public.du_lieu_gcn
# (không theo vị trí — cột "Phân loại thửa đất" nằm giữa mẫu (AY) chứ
# không cuối bảng như trong create_du_lieu_gcn.sql). None = cột mẫu không
# có dữ liệu tương ứng trong bảng (2 cột "Trạng thái xác thực số giấy tờ"
# BB/BE — chỉ có trên mẫu, không được đồng bộ/lưu).
_MAU_TANAN_COLUMNS = [
    ("B", "madvhc"),
    ("C", "sophathanhgcn"),
    ("D", "ngaycapgcn"),
    ("E", "sovaosogcn"),
    ("F", "tentochuc"),
    ("G", "madinhdanhtochuc"),
    ("H", "hovatenchusudung"),
    ("I", "sodinhdanhcanhan"),
    ("J", "ngaythangnamsinh"),
    ("K", "gioitinh"),
    ("L", "diachithuongtru"),
    ("M", "phapnhantrengcn"),
    ("N", "vaitrophapnhan"),
    ("O", "tenchusudunghientai"),
    ("P", "sodinhdanh_chusudunghientai"),
    ("Q", "diachi_chusudunghientai"),
    ("R", "lydothaydoi"),
    ("S", "madinhdanhthuadat"),
    ("T", "soto_gcn"),
    ("U", "sothua_gcn"),
    ("V", "soto"),
    ("W", "sothua"),
    ("X", "diachi_thuadat"),
    ("Y", "dientichthuadat"),
    ("Z", "loaidat1"),
    ("AA", "dientich1"),
    ("AB", "nguongoc1"),
    ("AC", "hinhthucsudung1"),
    ("AD", "thoihansudung1"),
    ("AE", "loaidat2"),
    ("AF", "dientich2"),
    ("AG", "nguongoc2"),
    ("AH", "hinhthucsudung2"),
    ("AI", "thoihansudung2"),
    ("AJ", "loaidat3"),
    ("AK", "dientich3"),
    ("AL", "nguongoc3"),
    ("AM", "hinhthucsudung3"),
    ("AN", "thoihansudung3"),
    ("AO", "loaitaisan"),
    ("AP", "khunhachungcu_honhop"),
    ("AQ", "nhachungcu"),
    ("AR", "socanho"),
    ("AS", "dientichxaydung"),
    ("AT", "dientichsan"),
    ("AU", "hinhthucsohuu"),
    ("AV", "thoihansohuu"),
    ("AW", "caphang"),
    ("AX", "tenfilequet"),
    ("AY", "phanloai"),
    ("AZ", "sogiayto1"),
    ("BA", "loaigiayto1"),
    ("BB", None),
    ("BC", "sogiayto2"),
    ("BD", "loaigiayto2"),
    ("BE", None),
    ("BF", "hosoquet"),
    ("BG", "guild"),
]

# Cột nên ghi dạng số (khớp kiểu số ở dữ liệu mẫu gốc) — dữ liệu trong
# du_lieu_gcn lưu dạng text nên chuyển đổi tốt nhất có thể, giữ nguyên text
# nếu không phải số hợp lệ.
_MAU_TANAN_INT_FIELDS = {"soto_gcn", "sothua_gcn", "soto", "sothua"}
_MAU_TANAN_FLOAT_FIELDS = {
    "dientichthuadat", "dientich1", "dientich2", "dientich3",
    "dientichxaydung", "dientichsan",
}


def _mau_tanan_cell_value(field: str, raw_value):
    value = (raw_value or "").strip() if isinstance(raw_value, str) else raw_value
    if value in (None, ""):
        return None
    if field in _MAU_TANAN_INT_FIELDS:
        try:
            return int(value)
        except (TypeError, ValueError):
            return value
    if field in _MAU_TANAN_FLOAT_FIELDS:
        try:
            return float(value)
        except (TypeError, ValueError):
            return value
    return value


def get_stats():
    result, error_response = gcn_repository.gcn_thu_thap_theo_xa()
    if error_response:
        return None, error_response

    rows = result if isinstance(result, list) else []
    items = []
    as_of = None
    for row in rows:
        ma_xa = row.get("ma_xa")
        if not ma_xa:
            continue
        tong = int(row.get("tong_so_thua") or 0)
        da_nhap = int(row.get("da_nhap_bieu") or 0)
        computed_at = row.get("computed_at")
        if computed_at and (as_of is None or computed_at > as_of):
            as_of = computed_at
        items.append(
            {
                "ma_xa": ma_xa,
                "tong_so_thua": tong,
                "da_nhap_bieu": da_nhap,
                "chua_nhap_bieu": max(tong - da_nhap, 0),
            }
        )
    return {"items": items, "as_of": as_of}, None


def refresh_stats_cache():
    result, error_response = gcn_repository.refresh_gcn_thu_thap_theo_xa_cache()
    if error_response:
        return None, error_response
    # Hàm SQL trả về số dòng đã ghi vào cache (PostgREST bọc trong scalar).
    rows = result if isinstance(result, int) else None
    return {"ok": True, "rows": rows}, None


def get_bieu_thong_ke():
    """Thống kê thửa đã nhập biểu theo xã, tách form (Nhóm 4) / nguồn
    khác, lấy TẤT CẢ thửa không phân biệt nhóm KH 2959 — khác get_stats()
    ở trên (chỉ tính nhóm 'chưa tạo lập dữ liệu')."""
    result, error_response = gcn_repository.bieu_thong_ke_theo_xa()
    if error_response:
        return None, error_response

    rows = result if isinstance(result, list) else []
    items = []
    as_of = None
    for row in rows:
        ma_xa = row.get("ma_xa")
        if not ma_xa:
            continue
        tong = int(row.get("tong_so_thua") or 0)
        tu_form = int(row.get("da_nhap_form") or 0)
        tu_nguon_khac = int(row.get("da_nhap_nguon_khac") or 0)
        da_nhap = int(row.get("da_nhap_bieu") or 0)
        computed_at = row.get("computed_at")
        if computed_at and (as_of is None or computed_at > as_of):
            as_of = computed_at
        items.append(
            {
                "ma_xa": ma_xa,
                "tong_so_thua": tong,
                "da_nhap_form": tu_form,
                "da_nhap_nguon_khac": tu_nguon_khac,
                "da_nhap_bieu": da_nhap,
                "chua_nhap_bieu": max(tong - da_nhap, 0),
            }
        )
    return {"items": items, "as_of": as_of}, None


def refresh_bieu_thong_ke_cache():
    result, error_response = gcn_repository.refresh_bieu_thong_ke_theo_xa_cache()
    if error_response:
        return None, error_response
    rows = result if isinstance(result, int) else None
    return {"ok": True, "rows": rows}, None


def export_bieu_thong_ke_xlsx():
    """Xuất .xlsx thống kê thửa đã nhập biểu theo xã (tách form/nguồn
    khác) — dựng từ dữ liệu cache sẵn có (get_bieu_thong_ke), không quét
    lại du_lieu_gcn nên rất nhẹ."""
    data, error_response = get_bieu_thong_ke()
    if error_response:
        return None, error_response

    xa_data, error_response = parcel_service.list_xa()
    ten_xa_by_ma = {}
    if not error_response:
        ten_xa_by_ma = {
            item["ma_xa"]: item.get("ten_xa") or "" for item in xa_data.get("items", [])
        }

    wb = Workbook()
    ws = wb.active
    ws.title = "ThongKeNhapBieu"[:31]
    ws.append([
        "Mã xã", "Tên xã/phường", "Tổng số thửa", "Từ form (Nhóm 4)",
        "Từ nguồn khác", "Tổng đã nhập biểu (duy nhất)", "Chưa nhập biểu",
        "Tỉ lệ đã nhập (%)",
    ])
    for row in data["items"]:
        tong = row["tong_so_thua"]
        percent = round((row["da_nhap_bieu"] / tong) * 100, 1) if tong else 0
        ws.append([
            row["ma_xa"],
            ten_xa_by_ma.get(row["ma_xa"], ""),
            tong,
            row["da_nhap_form"],
            row["da_nhap_nguon_khac"],
            row["da_nhap_bieu"],
            row["chua_nhap_bieu"],
            percent,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return (buf.getvalue(), "thong_ke_nhap_bieu_theo_xa.xlsx"), None


def _xlsx_cell(value):
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return str(value)
    return value


def export_theo_nhom_xlsx(ma_xa: str, nhom_raw: str | None):
    """Xuất .xlsx: các dòng du_lieu_gcn của xã đã thu thập GCN và thuộc
    nhóm KH 2959 chỉ định. Trả về ((bytes, ten_file), None) hoặc
    (None, response_loi)."""
    ma_xa = (ma_xa or "").strip()
    if not ma_xa:
        return None, (jsonify({"error": "Thiếu mã xã"}), 400)

    raw = nhom_raw or "Nhóm 2"
    nhom = [n.strip() for n in raw.split(",") if n.strip()]
    xau = [n.upper() for n in nhom if n.upper() in _NHOM_HOP_LE]
    if not xau:
        return None, (jsonify({"error": "Nhóm không hợp lệ (chỉ Nhóm 1 / Nhóm 2)"}), 400)

    rows: list = []
    offset = 0
    while True:
        page, error_response = gcn_repository.export_gcn_theo_nhom_page(
            ma_xa, nhom, _EXPORT_PAGE, offset
        )
        if error_response:
            return None, error_response
        page = page if isinstance(page, list) else []
        rows.extend(page)
        if len(page) < _EXPORT_PAGE or len(rows) >= _EXPORT_MAX:
            break
        offset += _EXPORT_PAGE

    wb = Workbook()
    ws = wb.active
    ws.title = "GCN"[:31]
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([_xlsx_cell(row.get(h)) for h in headers])
    else:
        ws.append(["Không có dữ liệu khớp"])

    buf = io.BytesIO()
    wb.save(buf)
    nhom_ten = "_".join(n.upper().replace("NHÓM ", "N").replace(" ", "") for n in nhom)
    return (buf.getvalue(), f"gcn_{nhom_ten}_{ma_xa}.xlsx"), None


def export_nhom3_xlsx(ma_xa: str, chi_nhom3: bool):
    """Xuất .xlsx đúng bố cục file mẫu TANAN.xlsx: toàn bộ dữ liệu GCN của
    1 đơn vị hành chính, hoặc chỉ thửa Nhóm 3 (chưa thuộc Nhóm 1/Nhóm 2 KH
    2959) nếu chi_nhom3=True. Trả về ((bytes, ten_file), None) hoặc
    (None, response_loi)."""
    ma_xa = (ma_xa or "").strip()
    if not ma_xa:
        return None, (jsonify({"error": "Thiếu mã đơn vị hành chính"}), 400)

    rows: list = []
    offset = 0
    while True:
        page, error_response = gcn_repository.export_du_lieu_gcn_page(
            ma_xa, chi_nhom3, _EXPORT_PAGE, offset
        )
        if error_response:
            return None, error_response
        page = page if isinstance(page, list) else []
        rows.extend(page)
        if len(page) < _EXPORT_PAGE or len(rows) >= _EXPORT_MAX:
            break
        offset += _EXPORT_PAGE

    try:
        wb = load_workbook(_MAU_TANAN_PATH)
    except FileNotFoundError:
        return None, (jsonify({"error": "Thiếu file mẫu TANAN.xlsx trên server"}), 500)
    ws = wb[_MAU_TANAN_SHEET] if _MAU_TANAN_SHEET in wb.sheetnames else wb.active

    # Xóa hết dữ liệu mẫu (hàng 5 trở đi) trước khi ghi dữ liệu thật.
    first_data_row = _MAU_TANAN_HEADER_ROWS + 1
    if ws.max_row >= first_data_row:
        ws.delete_rows(first_data_row, ws.max_row - first_data_row + 1)

    for i, row in enumerate(rows):
        r = first_data_row + i
        ws.cell(row=r, column=1, value=i + 1)  # cột A: Số TT
        for col_letter, field in _MAU_TANAN_COLUMNS:
            if field is None:
                continue
            ws[f"{col_letter}{r}"] = _mau_tanan_cell_value(field, row.get(field))

    buf = io.BytesIO()
    wb.save(buf)
    pham_vi = "nhom3" if chi_nhom3 else "toanbo"
    return (buf.getvalue(), f"du_lieu_gcn_{ma_xa}_{pham_vi}.xlsx"), None


def list_sources():
    result, error_response = gcn_repository.list_nguon()
    if error_response:
        return None, error_response
    return {"items": result}, None


def create_source(body: dict):
    ma_nguon = str(body.get("ma_nguon", "")).strip()
    ten_nguon = str(body.get("ten_nguon", "")).strip()
    url = str(body.get("url", "")).strip()
    kich_hoat = bool(body.get("kich_hoat", True))

    if not ma_nguon or not url:
        return None, (jsonify({"error": "Thiếu ma_nguon hoặc url"}), 400)

    payload = {"ma_nguon": ma_nguon, "ten_nguon": ten_nguon, "url": url, "kich_hoat": kich_hoat}
    result, error_response = gcn_repository.create_nguon(payload)
    if error_response:
        return None, error_response
    items = result
    return {"ok": True, "item": items[0] if items else payload}, None


def update_source(ma_nguon: str, body: dict):
    updates: dict = {}
    if "ten_nguon" in body:
        updates["ten_nguon"] = str(body["ten_nguon"]).strip()
    if "url" in body:
        updates["url"] = str(body["url"]).strip()
    if "kich_hoat" in body:
        updates["kich_hoat"] = bool(body["kich_hoat"])

    if not updates:
        return None, (jsonify({"error": "Không có trường nào để cập nhật"}), 400)

    updates["updated_at"] = datetime.now(timezone.utc).isoformat()

    result, error_response = gcn_repository.update_nguon(ma_nguon, updates)
    if error_response:
        return None, error_response
    if not result:
        return None, (jsonify({"error": "Không tìm thấy nguồn"}), 404)
    return {"ok": True, "item": result[0]}, None


def delete_source(ma_nguon: str):
    _, error_response = gcn_repository.delete_nguon(ma_nguon)
    if error_response:
        return None, error_response
    return {"ok": True}, None


def _get_supabase_context():
    base_url = supabase_client.get_base_url()
    if not base_url:
        return None, None, supabase_client.missing_base_url_response()
    try:
        headers = supabase_client.get_service_headers()
    except RuntimeError as exc:
        return None, None, (jsonify({"error": str(exc)}), 500)
    return base_url, headers, None


def sync_one(ma_nguon: str):
    base_url, headers, error_response = _get_supabase_context()
    if error_response:
        return None, error_response

    try:
        source, error_response = gcn_repository.fetch_nguon_row(ma_nguon)
    except requests.RequestException as exc:
        return None, (jsonify({"error": str(exc)}), 502)
    if error_response:
        return None, error_response
    if not source:
        return None, (jsonify({"error": "Không tìm thấy nguồn"}), 404)
    if not source.get("url"):
        return None, (jsonify({"error": "Nguồn chưa có URL"}), 400)

    try:
        imported = gcn_sync.sync_source(
            base_url, headers, ma_nguon, source.get("ten_nguon") or "", source["url"]
        )
    except RuntimeError as exc:
        return None, (jsonify({"error": str(exc)}), 500)
    except Exception as exc:  # noqa: BLE001 - lỗi đọc Sheet (sai URL/chưa share/sai tên trang tính...) phải báo rõ
        current_app.logger.exception("Đồng bộ GCN lỗi: ma_nguon=%s", ma_nguon)
        return None, (jsonify({"error": f"Đồng bộ thất bại: {exc}"}), 502)

    return {"ok": True, "ma_nguon": ma_nguon, "imported": imported}, None


def sync_all():
    base_url, headers, error_response = _get_supabase_context()
    if error_response:
        return None, error_response

    try:
        sources, error_response = gcn_repository.list_active_sources()
    except requests.RequestException as exc:
        return None, (jsonify({"error": str(exc)}), 502)
    if error_response:
        return None, error_response

    results = []
    for source in sources:
        ma_nguon = source.get("ma_nguon")
        url = source.get("url")
        if not ma_nguon or not url:
            continue

        try:
            imported = gcn_sync.sync_source(
                base_url, headers, ma_nguon, source.get("ten_nguon") or "", url
            )
        except Exception as exc:  # noqa: BLE001 - 1 nguồn lỗi không được chặn các nguồn khác
            current_app.logger.exception("Đồng bộ GCN lỗi: ma_nguon=%s", ma_nguon)
            results.append({"ma_nguon": ma_nguon, "ok": False, "error": str(exc)})
        else:
            results.append({"ma_nguon": ma_nguon, "ok": True, "imported": imported})

    return {"ok": True, "results": results}, None
