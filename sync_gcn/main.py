from __future__ import annotations

import sys
from dataclasses import dataclass
from datetime import datetime, timezone

import openpyxl

import google_sheets
import supabase_sync
from config import SHEET_LIST_FILE

REQUIRED_COLUMNS = ("ma_nguon", "ten_nguon", "url", "kich_hoat")


@dataclass
class SheetSource:
    ma_nguon: str
    ten_nguon: str
    url: str


def load_sources() -> list[SheetSource]:
    if not SHEET_LIST_FILE.exists():
        sys.exit(f"Không tìm thấy {SHEET_LIST_FILE}")

    workbook = openpyxl.load_workbook(SHEET_LIST_FILE, read_only=True, data_only=True)
    sheet = workbook.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    header = [str(value or "").strip().lower() for value in header_row]

    missing = [name for name in REQUIRED_COLUMNS if name not in header]
    if missing:
        sys.exit(f"{SHEET_LIST_FILE.name} thiếu cột bắt buộc: {', '.join(missing)}")

    col_index = {name: header.index(name) for name in REQUIRED_COLUMNS}

    sources: list[SheetSource] = []
    seen: set[str] = set()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row is None or all(value is None for value in row):
            continue

        ma_nguon = str(row[col_index["ma_nguon"]] or "").strip()
        ten_nguon = str(row[col_index["ten_nguon"]] or "").strip()
        url = str(row[col_index["url"]] or "").strip()
        kich_hoat = row[col_index["kich_hoat"]]

        if not ma_nguon or not url:
            continue
        if str(kich_hoat).strip() not in ("1", "1.0"):
            continue

        if ma_nguon in seen:
            sys.exit(f"ma_nguon trùng lặp trong {SHEET_LIST_FILE.name}: {ma_nguon}")
        seen.add(ma_nguon)

        sources.append(SheetSource(ma_nguon=ma_nguon, ten_nguon=ten_nguon, url=url))

    return sources


def sync_source(source: SheetSource) -> int:
    # 1) Đọc toàn bộ Google Sheet trước — nếu lỗi ở đây thì dừng lại
    #    ngay, chưa hề đụng tới Supabase.
    records = google_sheets.read_sheet_rows(source.url)

    # 2) Chuẩn hóa thành records hoàn chỉnh.
    synced_at = datetime.now(timezone.utc).isoformat()
    rows = []
    for record in records:
        row = dict(record)
        dong_sheet = row.pop("_dong_sheet")
        row["ma_nguon"] = source.ma_nguon
        row["ten_nguon"] = source.ten_nguon
        row["sheet_url"] = source.url
        row["dong_sheet"] = dong_sheet
        row["synced_at"] = synced_at
        rows.append(row)

    # 3) Chỉ xóa + ghi Supabase sau khi (1) và (2) đã xong an toàn.
    return supabase_sync.replace_source(source.ma_nguon, rows)


def main() -> None:
    sources = load_sources()
    if not sources:
        print(f"Không có nguồn nào kích hoạt (kich_hoat = 1) trong {SHEET_LIST_FILE.name}")
        return

    results: list[tuple[SheetSource, bool, int, str]] = []

    for source in sources:
        try:
            count = sync_source(source)
        except Exception as exc:  # noqa: BLE001 - lỗi 1 nguồn không được chặn các nguồn khác
            message = f"{type(exc).__name__}: {exc}"
            results.append((source, False, 0, message))
            print(f"[{source.ma_nguon}] {source.ten_nguon}: ERROR - {message}")
        else:
            results.append((source, True, count, ""))
            print(f"[{source.ma_nguon}] {source.ten_nguon}: OK - {count:,} dòng".replace(",", "."))

    ok = [r for r in results if r[1]]
    errors = [r for r in results if not r[1]]
    total_rows = sum(r[2] for r in ok)

    print()
    print("===== Tổng kết =====")
    print(f"Tổng nguồn  : {len(results)}")
    print(f"Thành công  : {len(ok)}")
    print(f"Lỗi         : {len(errors)}")
    print(f"Tổng bản ghi: {total_rows:,}".replace(",", "."))
    if errors:
        print("Nguồn lỗi:")
        for source, _, _, message in errors:
            print(f"  - [{source.ma_nguon}] {source.ten_nguon}: {message}")


if __name__ == "__main__":
    main()
