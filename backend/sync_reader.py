from __future__ import annotations

import csv
import io

from openpyxl import load_workbook

# Thứ tự đúng như các cột trong bảng dong_bo_du_lieu (supabase/schema.sql),
# lấy giá trị từ cột B đến cột O của file nguồn (cột A bị bỏ qua).
FIELDS = [
    "ma_xa",
    "so_to",
    "so_thua",
    "da_xuat_so_dia_chinh_dien_tu",
    "chua_xuat_so_dia_chinh_dien_tu",
    "dong_bo_3_khoi",
    "khong_dong_bo_3_khoi",
    "chi_co_du_lieu_thuoc_tinh",
    "khop_csdlqg_dan_cu",
    "chua_khop_csdlqg_dan_cu",
    "khong_xac_dinh_csdlqg_dan_cu",
    "van_hanh_24_7",
    "khong_van_hanh_24_7",
    "phan_loai_ke_hoach_2959",
]
INT_FIELDS = {"so_to", "so_thua"}
BOOL_FIELDS = {
    "da_xuat_so_dia_chinh_dien_tu",
    "chua_xuat_so_dia_chinh_dien_tu",
    "dong_bo_3_khoi",
    "khong_dong_bo_3_khoi",
    "chi_co_du_lieu_thuoc_tinh",
    "khop_csdlqg_dan_cu",
    "chua_khop_csdlqg_dan_cu",
    "khong_xac_dinh_csdlqg_dan_cu",
    "van_hanh_24_7",
    "khong_van_hanh_24_7",
}
COLUMN_START = 1  # cột B (0-based index)
COLUMN_END = COLUMN_START + len(FIELDS)  # cột O (exclusive)


def _to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _to_int(value) -> int:
    text = _to_text(value)
    return int(float(text)) if text else 0


def _to_bool(value) -> bool:
    return bool(_to_text(value))


def _row_to_record(cells: list) -> dict | None:
    values = list(cells[COLUMN_START:COLUMN_END])
    if len(values) < len(FIELDS):
        values += [""] * (len(FIELDS) - len(values))

    record = {}
    for field, value in zip(FIELDS, values):
        if field in INT_FIELDS:
            record[field] = _to_int(value)
        elif field in BOOL_FIELDS:
            record[field] = _to_bool(value)
        else:
            record[field] = _to_text(value) or None

    if not record["ma_xa"] or not record["so_to"] or not record["so_thua"]:
        return None
    return record


def _decode_csv_bytes(data: bytes) -> str:
    if data.startswith(b"\xff\xfe") or data.startswith(b"\xfe\xff"):
        return data.decode("utf-16")
    for encoding in ("utf-8-sig", "cp1258", "cp1252"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def parse_csv(data: bytes) -> list[dict]:
    text = _decode_csv_bytes(data)
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.reader(io.StringIO(text), dialect))
    return [record for row in rows[1:] if (record := _row_to_record(row))]


def parse_xlsx(data: bytes) -> list[dict]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    records = []
    for index, row in enumerate(sheet.iter_rows(values_only=True)):
        if index == 0:
            continue
        record = _row_to_record(list(row))
        if record:
            records.append(record)
    return records


def parse_sync_file(filename: str, data: bytes) -> list[dict]:
    lowered = filename.lower()
    if lowered.endswith(".csv"):
        return parse_csv(data)
    if lowered.endswith(".xlsx") or lowered.endswith(".xlsm"):
        return parse_xlsx(data)
    raise ValueError("Chỉ hỗ trợ file .csv hoặc .xlsx")
